"""
标注回写模块
读取用户在Key树形Excel中追加的两列标注，回写到ConfigMap数据文件的逐行Excel中
"""

import os
import re
from typing import Dict, List, Tuple, Any

import yaml
from openpyxl import load_workbook

from core.yaml_parser import YAMLParser, parse_properties_content
from utils.file_utils import ensure_dir


class AnnotationApplier:
    """标注回写器"""

    def __init__(self):
        self.parser = YAMLParser()
        self.errors = []

    # ------------------------------------------------------------------
    # 读取标注Excel → {full_key: (domain, component)}
    # 标注Excel列布局：Col1=领域, Col2=组件, Col3=层级1, Col4=层级2, ...
    # ------------------------------------------------------------------

    def load_annotated_excel(self, file_path: str) -> Dict[str, Tuple[str, str]]:
        """
        从标注Excel加载 key→(领域, 组件) 映射

        Returns:
            {full_dot_path: (domain, component)}
        """
        self.errors = []
        annotation_map: Dict[str, Tuple[str, str]] = {}

        if not file_path or not os.path.exists(file_path):
            self.errors.append(f"标注文件不存在: {file_path}")
            return annotation_map

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            self.errors.append(f"读取标注Excel失败: {str(e)}")
            return annotation_map

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return annotation_map

        # 路径栈：[(depth, segment)]，depth从1开始对应层级1
        path_stack: List[Tuple[int, str]] = []

        for row in rows[1:]:  # 跳过表头
            if not row:
                continue

            domain = str(row[0]).strip() if row[0] is not None else ""
            component = str(row[1]).strip() if row[1] is not None else ""

            # 找层级列（从col3开始，即index2）中最右非空值
            hierarchy_cols = row[2:]
            depth = None
            segment = None
            for i in range(len(hierarchy_cols) - 1, -1, -1):
                val = hierarchy_cols[i]
                if val is not None and str(val).strip():
                    depth = i + 1  # 1-based depth
                    segment = str(val).strip()
                    break

            if depth is None or segment is None:
                continue

            # 弹出深度 >= 当前深度的栈元素
            while path_stack and path_stack[-1][0] >= depth:
                path_stack.pop()
            path_stack.append((depth, segment))

            full_path = ".".join(seg for _, seg in path_stack)

            if domain or component:
                annotation_map[full_path] = (domain, component)

        try:
            wb.close()
        except Exception:
            pass

        return annotation_map

    # ------------------------------------------------------------------
    # 将标注应用到集群ConfigMap，导出Excel
    # ------------------------------------------------------------------

    def apply_annotations(
        self,
        cluster_path: str,
        annotation_map: Dict[str, Tuple[str, str]],
        output_dir: str,
    ) -> bool:
        """
        遍历集群中的ConfigMap资源，逐行标注并导出Excel

        输出结构：output_dir/namespace/ConfigMap/configmap_name/data_file.xlsx
        """
        from utils.excel_exporter import ExcelExporter

        self.errors = []

        if not os.path.exists(cluster_path):
            self.errors.append(f"集群路径不存在: {cluster_path}")
            return False

        if os.path.isfile(cluster_path):
            resources = self.parser.parse_yaml_file(cluster_path)
        else:
            resources = self.parser.parse_cluster_folder(cluster_path)

        configmap_resources = [r for r in resources if r.kind == "ConfigMap"]

        if not configmap_resources:
            self.errors.append("未找到ConfigMap类型资源")
            return False

        exporter = ExcelExporter()
        success = True

        for resource in configmap_resources:
            data: Dict[str, Any] = resource.yaml_content.get("data") or {}
            cm_name = resource.name
            namespace = resource.namespace

            for data_key, content in data.items():
                if not isinstance(content, str):
                    continue

                lines = self._annotate_data_file(data_key, content, annotation_map)

                # 构建输出路径
                out_dir = os.path.join(output_dir, namespace, "ConfigMap", cm_name)
                ensure_dir(out_dir)
                out_file = os.path.join(out_dir, f"{data_key}.xlsx")

                ok = exporter.export_annotated_config(lines, out_file)
                if not ok:
                    self.errors.extend(exporter.get_errors())
                    exporter.clear_errors()
                    success = False

        return success

    def _annotate_data_file(
        self,
        data_key: str,
        content: str,
        annotation_map: Dict[str, Tuple[str, str]],
    ) -> List[Tuple[str, str, str]]:
        """根据data key类型分发到对应的标注方法"""
        lower = data_key.lower()
        if lower.endswith(".yaml") or lower.endswith(".yml"):
            return self._annotate_yaml_lines(content, annotation_map)
        elif lower.endswith(".properties"):
            return self._annotate_properties_lines(content, annotation_map)
        else:
            # 纯文本，每行不标注
            return [("", "", line) for line in content.splitlines()]

    # ------------------------------------------------------------------
    # YAML逐行标注
    # ------------------------------------------------------------------

    def _annotate_yaml_lines(
        self,
        content: str,
        annotation_map: Dict[str, Tuple[str, str]],
    ) -> List[Tuple[str, str, str]]:
        """
        逐行扫描YAML内容，对叶子key行填写标注

        路径追踪：按缩进量维护路径栈，识别叶子节点
        """
        result: List[Tuple[str, str, str]] = []
        raw_lines = content.splitlines()

        # 路径栈：[(indent, key_segment)]
        path_stack: List[Tuple[int, str]] = []
        # 检测缩进单位（第一个缩进行的空格数）
        indent_unit = 2
        for line in raw_lines:
            stripped = line.lstrip()
            if stripped and line != stripped:
                indent_unit = len(line) - len(stripped)
                if indent_unit > 0:
                    break

        # 块标量跟踪：进入 | 或 > 后，后续行按缩进归属于该值
        in_block_scalar = False
        block_scalar_indent = -1

        for raw_line in raw_lines:
            stripped = raw_line.rstrip()

            # 计算缩进
            indent = len(stripped) - len(stripped.lstrip())

            # 块标量内部行：直接输出，不做路径解析
            if in_block_scalar:
                if stripped and indent > block_scalar_indent:
                    result.append(("", "", raw_line))
                    continue
                else:
                    in_block_scalar = False

            # 空行或注释行
            content_stripped = stripped.lstrip()
            if not content_stripped or content_stripped.startswith("#"):
                result.append(("", "", raw_line))
                continue

            # 列表项：简单忽略路径变化，直接输出
            if content_stripped.startswith("- "):
                result.append(("", "", raw_line))
                continue

            # 尝试匹配 key: value
            m = re.match(r'^(\s*)([\w\-. \'\"]+?)\s*:\s*(.*)$', stripped)
            if not m:
                result.append(("", "", raw_line))
                continue

            key_segment = m.group(2).strip().strip("'\"")
            value_part = m.group(3).strip()

            # 弹出缩进 >= 当前的路径栈
            while path_stack and path_stack[-1][0] >= indent:
                path_stack.pop()

            # 判断是否叶子节点
            is_leaf = False
            is_block = False

            if value_part in ("|", ">", "|-", ">-", "|+", ">+"):
                is_leaf = True
                is_block = True
            elif value_part and value_part not in ("{}", "[]"):
                is_leaf = True
            # else: value为空，是容器节点

            # 更新路径栈（无论叶子还是容器都入栈，容器的子节点需要它）
            path_stack.append((indent, key_segment))

            full_path = ".".join(seg for _, seg in path_stack)

            if is_leaf:
                annotation = annotation_map.get(full_path, ("", ""))
                result.append((annotation[0], annotation[1], raw_line))
                if is_block:
                    in_block_scalar = True
                    block_scalar_indent = indent
            else:
                result.append(("", "", raw_line))

        return result

    # ------------------------------------------------------------------
    # Properties逐行标注
    # ------------------------------------------------------------------

    def _annotate_properties_lines(
        self,
        content: str,
        annotation_map: Dict[str, Tuple[str, str]],
    ) -> List[Tuple[str, str, str]]:
        """逐行扫描properties内容，对存在标注的key填写标注"""
        result: List[Tuple[str, str, str]] = []
        for raw_line in content.splitlines():
            stripped = raw_line.strip()
            if not stripped or stripped.startswith("#") or stripped.startswith("!"):
                result.append(("", "", raw_line))
                continue

            sep_idx = -1
            for i, ch in enumerate(stripped):
                if ch in ("=", ":"):
                    sep_idx = i
                    break

            if sep_idx > 0:
                key = stripped[:sep_idx].strip()
                annotation = annotation_map.get(key, ("", ""))
                result.append((annotation[0], annotation[1], raw_line))
            else:
                result.append(("", "", raw_line))

        return result

    # ------------------------------------------------------------------
    # 错误处理
    # ------------------------------------------------------------------

    def get_errors(self) -> List[str]:
        return self.errors + self.parser.get_errors()

    def clear_errors(self):
        self.errors = []
        self.parser.clear_errors()
