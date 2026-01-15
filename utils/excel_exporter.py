"""
Excel导出工具
负责将数据导出为Excel文件
"""

import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from models.resource import Resource, ComparisonResult, ExtractionResult
from config import MAX_SHEETS_PER_FILE


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        self.errors = []
    
    def export_resource_table(self, resources: List[Resource], output_path: str) -> bool:
        """
        导出资源表格
        
        Args:
            resources: 资源列表
            output_path: 输出文件路径
            
        Returns:
            是否成功
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "资源列表"
            
            # 设置表头
            headers = ["集群名", "命名空间", "资源类型", "资源名", "文件路径"]
            self._write_header(ws, headers)
            
            # 写入数据
            for idx, resource in enumerate(resources, start=2):
                ws.cell(row=idx, column=1, value=resource.cluster)
                ws.cell(row=idx, column=2, value=resource.namespace)
                ws.cell(row=idx, column=3, value=resource.kind)
                ws.cell(row=idx, column=4, value=resource.name)
                ws.cell(row=idx, column=5, value=resource.file_path)
            
            # 自动调整列宽
            self._adjust_column_width(ws)
            
            # 保存文件
            wb.save(output_path)
            return True
            
        except Exception as e:
            self.errors.append(f"导出资源表失败: {str(e)}")
            return False
    
    def export_comparison_result(self, comparison_results: List[ComparisonResult], 
                                 output_dir: str, cluster1_name: str, cluster2_name: str) -> bool:
        """
        导出比较结果
        按资源类型分文件，每个文件按资源名分Sheet
        
        Args:
            comparison_results: 比较结果列表
            output_dir: 输出目录
            cluster1_name: 集群1名称
            cluster2_name: 集群2名称
            
        Returns:
            是否成功
        """
        try:
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 按资源类型分组
            grouped_by_kind = {}
            for result in comparison_results:
                kind = result.resource_left.kind
                if kind not in grouped_by_kind:
                    grouped_by_kind[kind] = []
                grouped_by_kind[kind].append(result)
            
            # 为每个资源类型创建Excel文件
            for kind, results in grouped_by_kind.items():
                self._export_comparison_by_kind(
                    kind, results, output_dir, cluster1_name, cluster2_name
                )
            
            return True
            
        except Exception as e:
            self.errors.append(f"导出比较结果失败: {str(e)}")
            return False
    
    def _export_comparison_by_kind(self, kind: str, results: List[ComparisonResult], 
                                   output_dir: str, cluster1_name: str, cluster2_name: str):
        """导出单个资源类型的比较结果"""
        # 按资源名分组
        grouped_by_name = {}
        for result in results:
            name = result.resource_left.name
            if name not in grouped_by_name:
                grouped_by_name[name] = []
            grouped_by_name[name].append(result)
        
        # 处理Sheet数量限制
        resource_names = list(grouped_by_name.keys())
        file_count = 1
        start_idx = 0
        
        while start_idx < len(resource_names):
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认sheet
            
            end_idx = min(start_idx + MAX_SHEETS_PER_FILE, len(resource_names))
            
            for name in resource_names[start_idx:end_idx]:
                results_for_name = grouped_by_name[name]
                ws = wb.create_sheet(title=self._sanitize_sheet_name(name))
                
                # 写入表头
                headers = ["比较Key路径", f"{cluster1_name}值", f"{cluster2_name}值", "差异标记"]
                self._write_header(ws, headers)
                
                # 写入差异数据
                row_idx = 2
                for result in results_for_name:
                    if result.resource_right is None:
                        # 右侧资源不存在
                        ws.cell(row=row_idx, column=1, value="[整个资源]")
                        ws.cell(row=row_idx, column=2, value="存在")
                        ws.cell(row=row_idx, column=3, value="不存在")
                        ws.cell(row=row_idx, column=4, value="缺失")
                        self._highlight_row(ws, row_idx, "FFCCCC")
                        row_idx += 1
                    else:
                        for diff in result.differences:
                            ws.cell(row=row_idx, column=1, value=diff['key_path'])
                            ws.cell(row=row_idx, column=2, value=str(diff['left_value']))
                            ws.cell(row=row_idx, column=3, value=str(diff['right_value']))
                            ws.cell(row=row_idx, column=4, value="不同")
                            self._highlight_row(ws, row_idx, "FFFFCC")
                            row_idx += 1
                
                self._adjust_column_width(ws)
            
            # 保存文件
            if len(resource_names) > MAX_SHEETS_PER_FILE:
                filename = f"{kind}_比较结果_{file_count}.xlsx"
            else:
                filename = f"{kind}_比较结果.xlsx"
            
            output_path = os.path.join(output_dir, filename)
            wb.save(output_path)
            
            start_idx = end_idx
            file_count += 1
    
    def export_extraction_result(self, extraction_results: List[ExtractionResult], 
                                 output_dir: str) -> bool:
        """
        导出信息提取结果
        按资源类型分文件，每个文件为行式表格
        
        Args:
            extraction_results: 提取结果列表
            output_dir: 输出目录
            
        Returns:
            是否成功
        """
        try:
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 按资源类型分组
            grouped_by_kind = {}
            for result in extraction_results:
                kind = result.resource.kind
                if kind not in grouped_by_kind:
                    grouped_by_kind[kind] = []
                grouped_by_kind[kind].append(result)
            
            # 为每个资源类型创建Excel文件
            for kind, results in grouped_by_kind.items():
                self._export_extraction_by_kind(kind, results, output_dir)
            
            return True
            
        except Exception as e:
            self.errors.append(f"导出提取结果失败: {str(e)}")
            return False
    
    def _export_extraction_by_kind(self, kind: str, results: List[ExtractionResult], 
                                   output_dir: str):
        """导出单个资源类型的提取结果"""
        wb = Workbook()
        ws = wb.active
        ws.title = "提取结果"

        # 收集所有key列
        key_columns = []
        for result in results:
            for key in result.extracted_values.keys():
                if key not in key_columns:
                    key_columns.append(key)

        headers = ["集群名", "命名空间", "资源名"] + key_columns
        self._write_header(ws, headers)

        row_idx = 2
        for result in results:
            ws.cell(row=row_idx, column=1, value=result.resource.cluster)
            ws.cell(row=row_idx, column=2, value=result.resource.namespace)
            ws.cell(row=row_idx, column=3, value=result.resource.name)

            for col_offset, key in enumerate(key_columns, start=4):
                value = result.extracted_values.get(key, "")
                ws.cell(row=row_idx, column=col_offset, value=str(value) if value is not None else "")
            row_idx += 1

        self._adjust_column_width(ws)

        filename = f"{kind}_提取结果.xlsx"
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)
    
    def _write_header(self, ws, headers: List[str]):
        """写入表头并设置样式"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _highlight_row(self, ws, row_idx: int, color: str):
        """高亮整行"""
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    
    def _adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """清理Sheet名称，移除非法字符"""
        # Excel Sheet名称不能包含: : \ / ? * [ ]
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # 限制长度（Excel限制为31个字符）
        return name[:31]
    
    def get_errors(self) -> List[str]:
        """获取错误列表"""
        return self.errors
    
    def clear_errors(self):
        """清除错误列表"""
        self.errors = []
